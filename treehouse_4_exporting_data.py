from treehouse_3_cleaning_data import *

def write_to_file(filename, sample_data):
	to_csv = csv.writer(open(filename, 'w', encoding='utf-8'), dialect='excel')
	to_csv.writerows(sample_data)

#write_to_file('s4_gucci_ties.csv', gucci_ties)

def write_brand_and_price_csv(filename, sample_data):
	brand_col = 5
	price_col = 2

	the_list = []
	for record in sample_data:
		new_record = []
		new_record.append(record[brand_col])
		new_record.append(record[price_col])
		the_list.append(new_record)

	write_to_file(filename, the_list)


#write_brand_and_price_csv('s4_brand_and_price.csv', jcrew_ties)

def write_min_max_csv(filename, sample_data):
	the_max = find_max_min(jcrew_ties[1:], 2, 'max')
	the_min = find_max_min(jcrew_ties[1:], 2, 'min')

	the_list = []
	for record in sample_data[1:]:
		if (float(record[2]) == the_max) or (float(record[2]) == the_min):
			the_list.append(record)

	write_to_file(filename, the_list)


write_min_max_csv('s4_min_max_price.csv', jcrew_ties)

def write_two_cols_csv(filename, sample_data, col1=0, col2=1):
	the_list = []
	for record in sample_data:
		new_record = []
		new_record.append(record[col1])
		new_record.append(record[col2])
		the_list.append(new_record)

	write_to_file(filename, the_list)

#write_two_cols_csv('s4_two_cols.csv', jcrew_ties, 3, 7)

#csv with sorted prices
# ascending is going up i.e. A-Z 
# descending is going down i.e. Z-A 

def write_sorted_prices(filename, data_sample, order="ascending"):

    if order == "descending":
        data_sample.sort(key=lambda x: float(x[2]), reverse=False)
    else:
        data_sample.sort(key=lambda x: float(x[2]), reverse=True)

    write_to_file(filename, data_sample) 

write_sorted_prices('s4_sorted_prices.csv', striped_ties[1:], "descending")


#append another file
def write_append_file(filename, new_data_to_add):
    with open(filename, "a", encoding='utf-8') as myfile:
        for row in new_data_to_add:
            myfile.write(str(row))


import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter


def write_to_excel(filename, sample_data):
	the_workbook = Workbook()		#create new workbook
	active_sheet = the_workbook.active 	#create active spreadsheet inside workbook

	row_index = 1
	for record in sample_data:	#loop through every row
		col_index = 1
		for field in record:	#loop through every field/column in a row
			col_letter = get_column_letter(col_index) #get the letter name for each field (A or B)
			active_sheet.cell("{}{}".format(col_letter, row_index)).value = field
			#write the field value to the cell

			col_index += 1
		row_index +=1

	the_workbook.save(filename)

kiton_ties = filter_col_by_string(data_csv, 'brandName', 'Kiton')
#write_to_excel('s4_write_to_excel.xlsx', kiton_ties)

